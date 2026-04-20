"""
Sprint Retrospective Excel Generator
AI-Powered Hospital Management System
Format: matches the 4-column Sprint Retrospective board seen in the reference image
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def border_all(color="AAAAAA", style="thin"):
    s = Side(style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def border_bottom(color="AAAAAA"):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def apply(cell, value="", bold=False, italic=False, size=10,
          font_color="000000", bg=None, h_align="left",
          v_align="center", wrap=True, border=True):
    cell.value = value
    cell.font = Font(bold=bold, italic=italic, size=size,
                     color=font_color, name="Calibri")
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align,
                                wrap_text=wrap)
    if bg:
        cell.fill = fill(bg)
    if border:
        cell.border = border_all()

# ──────────────────────────────────────────────────────────
# SHEET BUILDER
# Sprint Retrospective 4-column board matching the reference image
# Columns: What Went Well | What Went Poorly | What Ideas Do You Have | How Should We Take Action
# Plus a narrow "Guidelines" column (E) that mirrors the ref image
# ──────────────────────────────────────────────────────────
def build_retro_sheet(wb, sprint_number, content):
    ws = wb.create_sheet(f"Sprint {sprint_number} Retro")
    ws.sheet_view.showGridLines = False

    # Column widths  A=narrow index | B C D E = four main cols | F = guideline col
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 12

    # ── ROW 1: Main title banner ──────────────────────────────
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Sprint Retrospective"
    t.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = fill("1F6AA5")      # dark blue matching image
    t.border = border_all("1F6AA5")
    ws.row_dimensions[1].height = 32

    # ── ROW 2: Sprint meta info ───────────────────────────────
    ws.merge_cells("A2:F2")
    m = ws["A2"]
    m.value = (f"Project: AI-Powered Hospital Management System  |  "
               f"Sprint {sprint_number}  |  "
               f"Duration: {content['duration']}  |  "
               f"Date: {content['date']}")
    m.font = Font(bold=False, size=9, color="FFFFFF", name="Calibri", italic=True)
    m.alignment = Alignment(horizontal="center", vertical="center")
    m.fill = fill("2E86C1")
    m.border = border_all("2E86C1")
    ws.row_dimensions[2].height = 18

    # ── ROW 3: Column headers ─────────────────────────────────
    HEADER_BG   = "2980B9"   # medium blue
    HEADER_FONT = "FFFFFF"
    headers = [
        ("B3", "What Went Well ✅"),
        ("C3", "What Went Poorly ❌"),
        ("D3", "What Ideas Do You Have 💡"),
        ("E3", "How Should We Take Action 🚀"),
    ]
    ws.row_dimensions[3].height = 30
    ws["A3"].fill = fill(HEADER_BG)
    ws["A3"].border = border_all(HEADER_BG)
    ws["F3"].value = "Guidelines"
    ws["F3"].font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
    ws["F3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["F3"].fill = fill(HEADER_BG)
    ws["F3"].border = border_all(HEADER_BG)

    for cell_id, label in headers:
        c = ws[cell_id]
        c.value = label
        c.font = Font(bold=True, size=11, color=HEADER_FONT, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = fill(HEADER_BG)
        c.border = border_all("1A6B9A")

    # ── ROW 4: Guideline descriptions (italic, light bg) ──────
    GUIDE_BG = "D6EAF8"
    guide_texts = [
        ("B4", ("This section highlights the successes and positive outcomes "
                "from the sprint. It helps the team recognize achievements "
                "and identify practices that should be continued.")),
        ("C4", ("This section identifies the challenges, roadblocks, or "
                "failures encountered during the sprint. It helps pinpoint "
                "areas that need improvement or change.")),
        ("D4", ("This section is for brainstorming new approaches, tools, "
                "or strategies to enhance the team's efficiency, "
                "productivity, or project outcomes.")),
        ("E4", ("This section outlines specific steps or solutions to address "
                "the issues and implement the ideas discussed, ensuring "
                "continuous improvement in future sprints.")),
    ]
    ws.row_dimensions[4].height = 70
    ws["A4"].fill = fill(GUIDE_BG)
    ws["A4"].border = border_all()

    guideline_col = ws["F4"]
    guideline_col.value = "Guidelines"
    guideline_col.font = Font(italic=True, size=8, color="555555", name="Calibri")
    guideline_col.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    guideline_col.fill = fill("EBF5FB")
    guideline_col.border = border_all()

    for cell_id, txt in guide_texts:
        c = ws[cell_id]
        c.value = txt
        c.font = Font(italic=True, size=9, color="1A5276", name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.fill = fill(GUIDE_BG)
        c.border = border_all("A9CCE3")

    # ── DATA ROWS ─────────────────────────────────────────────
    # Zip the four columns, padding shorter lists with ""
    went_well  = content["went_well"]
    went_bad   = content["went_bad"]
    ideas      = content["ideas"]
    actions    = content["actions"]

    max_rows = max(len(went_well), len(went_bad), len(ideas), len(actions))
    # Pad
    for lst in [went_well, went_bad, ideas, actions]:
        while len(lst) < max_rows:
            lst.append("")

    ALT1 = "FDFEFE"   # near-white
    ALT2 = "EBF5FB"   # very light blue

    for i in range(max_rows):
        row = 5 + i
        bg  = ALT1 if i % 2 == 0 else ALT2
        ws.row_dimensions[row].height = 48

        # Index cell (A)
        idx = ws.cell(row, 1, value=i + 1)
        idx.font = Font(size=9, color="888888", name="Calibri")
        idx.alignment = Alignment(horizontal="center", vertical="center")
        idx.fill = fill("F0F3F4")
        idx.border = border_all()

        pairs = [
            (2, went_well[i]),
            (3, went_bad[i]),
            (4, ideas[i]),
            (5, actions[i]),
        ]
        for col, val in pairs:
            c = ws.cell(row, col, value=val)
            c.font = Font(size=10, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.fill = fill(bg)
            c.border = border_all()

        # Guideline column stays blank in data rows
        g = ws.cell(row, 6, value="")
        g.fill = fill("F2F3F4")
        g.border = border_all()

    # ── Empty buffer rows at the bottom (like in the image) ───
    for extra in range(5):
        row = 5 + max_rows + extra
        ws.row_dimensions[row].height = 32
        for col in range(1, 7):
            c = ws.cell(row, col, value="")
            c.fill = fill("FFFFFF")
            c.border = border_all("CCCCCC")

    return ws


# ──────────────────────────────────────────────────────────
# SHEET 2: Full Narrative Retrospective Report
# ──────────────────────────────────────────────────────────
def build_report_sheet(wb, content):
    ws = wb.create_sheet("Full Report")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    ws.column_dimensions["C"].width = 3

    def add_row(value="", bold=False, italic=False, size=10,
                font_color="000000", bg="FFFFFF", h="auto",
                h_align="left", row_h=None):
        ws.append([" ", value, " "])
        r = ws.max_row
        c = ws.cell(r, 2)
        c.font = Font(bold=bold, italic=italic, size=size,
                      color=font_color, name="Calibri")
        c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
        c.fill = fill(bg)
        if row_h:
            ws.row_dimensions[r].height = row_h
        for col in [1, 3]:
            ws.cell(r, col).fill = fill(bg)
        return c

    # Title
    add_row("Sprint Retrospective Report", bold=True, size=18,
            font_color="FFFFFF", bg="1F6AA5", h_align="center", row_h=40)
    add_row("AI-Powered Hospital Management System (React + FastAPI + Supabase + ML)",
            italic=True, size=11, font_color="FFFFFF", bg="2E86C1",
            h_align="center", row_h=22)
    add_row(f"Sprint {content['sprint_no']}  |  {content['duration']}  |  {content['date']}",
            italic=True, size=9, font_color="FFFFFF", bg="2980B9",
            h_align="center", row_h=18)
    add_row(row_h=8)  # spacer

    SECTION_BG   = "D6EAF8"
    SECTION_FONT = "1A5276"

    def section(title):
        add_row(row_h=6)
        c = add_row(title, bold=True, size=12, font_color=SECTION_FONT,
                    bg=SECTION_BG, row_h=28)

    def bullet(text, sub=False):
        prefix = "      •  " if sub else "  •  "
        add_row(prefix + text, size=10, bg="FFFFFF", row_h=20)

    def para(text):
        add_row(text, size=10, bg="FDFEFE", row_h=None)

    # ── 1. Sprint Overview ────────────────────────────────────
    section("1.  Sprint Overview")
    bullet(f"Sprint Number   :  {content['sprint_no']}")
    bullet(f"Duration        :  {content['duration']}")
    bullet(f"Reporting Date  :  {content['date']}")
    bullet("Team            :  Full-stack development team (React + FastAPI + ML)")
    bullet("Sprint Goal     :")
    for g in content["goals"]:
        bullet(g, sub=True)

    # ── 2. What Went Well ─────────────────────────────────────
    section("2.  What Went Well  ✅")
    for item in content["went_well"]:
        if item:
            bullet(item)

    # ── 3. What Didn't Go Well ────────────────────────────────
    section("3.  What Didn't Go Well  ❌")
    for item in content["went_bad"]:
        if item:
            bullet(item)

    # ── 4. What Can Be Improved ───────────────────────────────
    section("4.  What Can Be Improved  🔧")
    for item in content["ideas"]:
        if item:
            bullet(item)

    # ── 5. Action Items ───────────────────────────────────────
    section("5.  Action Items for Next Sprint  🚀")
    for item in content["actions"]:
        if item:
            bullet(item)

    # ── 6. Key Learnings ─────────────────────────────────────
    section("6.  Key Learnings  🧠")
    for item in content["learnings"]:
        bullet(item)

    # ── 7. Tools & Technologies ───────────────────────────────
    section("7.  Tools & Technologies Used")
    for item in content["tools"]:
        bullet(item)

    add_row(row_h=12)
    add_row("— End of Sprint Retrospective Report —",
            italic=True, size=9, font_color="AAAAAA",
            h_align="center", row_h=20)

    return ws


# ──────────────────────────────────────────────────────────
# CONTENT
# ──────────────────────────────────────────────────────────
content = {
    "sprint_no" : "Sprint 1",
    "duration"  : "3 Weeks  (06 Mar 2026 – 27 Mar 2026)",
    "date"      : "27 March 2026",

    "goals": [
        "Build core Hospital Management System features (Patient, Room, Equipment modules)",
        "Integrate ML models — Random Forest, XGBoost, Logistic Regression, Prophet",
        "Develop AI Insights dashboard with Length-of-Stay, Readmission & Forecast predictions",
        "Add model evaluation metrics (accuracy, AUC, RMSE) and Explainable AI (feature importance)",
        "Establish Supabase backend with synthetic data pipeline and FastAPI REST layer",
    ],

    "went_well": [
        "✅  Full-stack integration (React ↔ FastAPI ↔ Supabase) delivered end-to-end data flow within the first week.",
        "✅  ML models (Random Forest, XGBoost, Logistic Regression) trained and deployed via FastAPI; predictions visible in the UI.",
        "✅  Prophet model successfully integrated for patient arrival forecasting with trend & seasonality decomposition.",
        "✅  Interactive dashboards built with Recharts — KPI cards, sparklines, and trend charts rendered from live API data.",
        "✅  Explainable AI layer implemented with feature-importance proxies surfaced in the AI Insights page.",
        "✅  Parallel API calls (Promise.all) refactored across all pages, eliminating sequential load-time bottlenecks.",
        "✅  Supabase feedback loop established to capture prediction feedback for future model retraining.",
        "✅  Task division among team members was clear; daily stand-up notes kept everyone aligned.",
        "✅  CORS, authentication headers, and environment variables configured correctly from day one, avoiding late-stage deployment issues.",
    ],

    "went_bad": [
        "❌  Real-world healthcare data was unavailable; synthetic data introduced unrealistic patterns that affected model generalisation.",
        "❌  XGBoost hyperparameter tuning was time-consuming; grid-search runs locked the notebook kernel for extended periods.",
        "❌  Frontend–ML API integration issues arose due to mismatched JSON response schemas (snake_case vs camelCase).",
        "❌  Full SHAP explainability (SHAP TreeExplainer) could not be completed within the sprint due to library version conflicts and time constraints.",
        "❌  Debugging FastAPI dependency injection errors (missing Pydantic validators) cost nearly a full day.",
        "❌  Supabase RLS (Row-Level Security) policies caused unexpected 403 errors during initial integration testing.",
        "❌  Prophet model required significant pre-processing (datetime indexing, frequency alignment) that was underestimated in planning.",
        "❌  UI responsiveness on smaller screen sizes was inconsistent — media queries were incomplete at the end of the sprint.",
    ],

    "ideas": [
        "💡  Source a real, anonymised hospital dataset (e.g., MIMIC-III) to improve model validity and academic credibility.",
        "💡  Implement full SHAP explainability using TreeExplainer for Random Forest and XGBoost models.",
        "💡  Add Optuna-based automated hyperparameter optimisation to reduce manual tuning overhead.",
        "💡  Standardise API response schemas with a shared Pydantic base model and a camelCase serialiser on the FastAPI side.",
        "💡  Introduce a CI/CD pipeline (GitHub Actions) to auto-deploy backend and run regression tests on each push.",
        "💡  Add real-time data update support using Supabase Realtime subscriptions instead of polling.",
        "💡  Create a high-risk patient alert system that triggers notifications when model confidence exceeds a threshold.",
        "💡  Refactor frontend into smaller, reusable components to improve maintainability and test coverage.",
    ],

    "actions": [
        "🚀  Improve ML model accuracy: source better data, add cross-validation, and document evaluation metrics formally.",
        "🚀  Complete SHAP integration — resolve library conflicts and add SHAP force plots to the AI Insights dashboard.",
        "🚀  Standardise API contracts: create shared Pydantic schemas and update frontend API service layer accordingly.",
        "🚀  Implement high-risk patient alerts with email/in-app notifications using Supabase Edge Functions.",
        "🚀  Expand dashboard analytics: add cohort analysis, bed-occupancy heatmap, and equipment failure trend charts.",
        "🚀  Add authentication security improvements: JWT token refresh, role-based access control (Admin / Doctor / Nurse).",
        "🚀  Write unit & integration tests for explainability.py, patient_models.py, and all FastAPI endpoints.",
        "🚀  Fix UI responsiveness issues — apply consistent CSS Grid / Flexbox breakpoints across all pages.",
        "🚀  Document the ML pipeline, data schema, and API contracts in a project Wiki / README.",
    ],

    "learnings": [
        "🧠  Clean, well-labelled data is the single most important prerequisite for reliable ML predictions — model quality is bounded by data quality.",
        "🧠  Deploying ML models in a web context requires careful API contract design; schema mismatches cause cascading frontend failures.",
        "🧠  Full-stack integration complexity grows non-linearly — allocate explicit integration buffer time in sprint planning.",
        "🧠  Explainability (XAI) is not optional in healthcare AI; stakeholders need to understand why a prediction was made, not just what it is.",
        "🧠  Performance optimisation (parallel fetching, caching) should be part of the initial architecture, not an afterthought.",
        "🧠  User-friendly UI in healthcare is critical — a clinician losing trust in the interface will stop using the tool regardless of model accuracy.",
        "🧠  Version-pinning ML dependencies (scikit-learn, SHAP, XGBoost) from the start prevents painful conflicts in later sprints.",
    ],

    "tools": [
        "Frontend  :  React 18, Recharts, Chart.js, CoreUI Pro (design inspiration), CSS Flexbox / Grid",
        "Backend   :  FastAPI (Python 3.11), Pydantic v2, Uvicorn, Python-JOSE (JWT)",
        "Database  :  Supabase (PostgreSQL), Supabase Realtime, Row-Level Security",
        "ML / AI   :  Scikit-learn (Random Forest, Logistic Regression, K-Fold CV), XGBoost, Prophet (Meta), SHAP (partial)",
        "Dev Tools :  VS Code, Jupyter Notebook, Git, GitHub, Postman, Python venv",
        "Deployment:  Local (dev), planned: Vercel (frontend) + Railway / Render (FastAPI backend)",
    ],
}

# ──────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ──────────────────────────────────────────────────────────
wb.remove(wb.active)   # remove default blank sheet

build_retro_sheet(wb, "1", content)
build_report_sheet(wb, content)

output = r"d:\Minor Project goated\Sprint_Retrospective.xlsx"
wb.save(output)
print(f"✅  Done!  Saved to: {output}")
