"""
MS Planner — 4 Bucket Board Generator
AI-Powered Hospital Management System
Buckets: Product Backlog | Sprint Backlog | In Progress | Testing/Completed
Includes: Concise checklists per major task
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def side(color="CCCCCC", style="thin"):
    return Side(style=style, color=color)

def border(color="CCCCCC"):
    s = side(color)
    return Border(top=s, bottom=s, left=s, right=s)

def cell(ws, row, col, value="", bold=False, italic=False, size=10,
         fc="1A1A2E", bg=None, ha="left", va="center", wrap=True, b=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, italic=italic, size=size, color=fc, name="Segoe UI")
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bg:
        c.fill = fill(bg)
    if b:
        c.border = border()
    return c

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR PALETTE  (MS Planner-inspired)
# ─────────────────────────────────────────────────────────────────────────────
DARK_BG     = "1B1F3B"   # deep navy — title bar
BUCKET_CLR  = {          # bucket accent colours
    "Product Backlog":      ("5C6BC0", "EDE7F6"),   # indigo
    "Sprint Backlog":       ("F57C00", "FFF3E0"),   # amber
    "In Progress":          ("1976D2", "E3F2FD"),   # blue
    "Testing/Completed":    ("388E3C", "E8F5E9"),   # green
}

LABEL_COLORS = {
    "Frontend":      ("E3F2FD", "1565C0"),
    "Backend":       ("FFF9C4", "F57F17"),
    "ML/AI":         ("FCE4EC", "B71C1C"),
    "Database":      ("E8F5E9", "1B5E20"),
    "DevOps":        ("EDE7F6", "4527A0"),
    "Security":      ("FBE9E7", "BF360C"),
    "Testing":       ("F3E5F5", "6A1B9A"),
    "Documentation": ("E0F2F1", "004D40"),
}

PRIORITY = {
    "Critical": ("FF1744", "FFFFFF"),
    "High":     ("FF6D00", "FFFFFF"),
    "Medium":   ("FFD600", "1A1A2E"),
    "Low":      ("00C853", "FFFFFF"),
}

# ─────────────────────────────────────────────────────────────────────────────
# TASK DATA (Major items only, concise checklists)
# Fields: (task_name, assignee, label, priority, due_date, notes, checklist)
# ─────────────────────────────────────────────────────────────────────────────
BUCKETS = {
    "Product Backlog": [
        (
            "Implement SHAP TreeExplainer (full XAI)",
            "Backend Dev", "ML/AI", "High", "Apr 15",
            "Replace proxy XAI with real SHAP force plots",
            [
                "☐ Replace proxy explainability with shap.TreeExplainer(model)",
                "☐ Serialise SHAP output as JSON for frontend"
            ]
        ),
        (
            "Role-based access control (Admin/Doctor/Nurse)",
            "Full Stack", "Security", "High", "Apr 20",
            "Supabase RLS + FastAPI roles",
            [
                "☐ Add role column to users metadata",
                "☐ Write RLS policies for Admin, Doctor, Nurse"
            ]
        ),
        (
            "Real-time updates via Supabase Realtime",
            "Frontend Dev", "Frontend", "Medium", "Apr 25",
            "Replace polling with ws subscriptions",
            [
                "☐ Remove existing polling from Patient & Room pages",
                "☐ Set up supabase.channel() subscription for state updates"
            ]
        ),
        (
            "CI/CD pipeline with GitHub Actions",
            "DevOps", "DevOps", "Medium", "Apr 28",
            "Auto-deploy FastAPI to Render + Vite to Vercel",
            [
                "☐ Create deploy workflow for backend",
                "☐ Create deploy workflow for frontend"
            ]
        )
    ],

    "Sprint Backlog": [
        (
            "Fix patient_models.py heuristic bug",
            "Backend Dev", "Backend", "Critical", "Mar 30",
            "Models skipping predictions",
            [
                "☐ Remove `pass` statements in try blocks",
                "☐ Verify fallback logic behavior"
            ]
        ),
        (
            "Pytest suite for hospital_predictions.py",
            "Backend Dev", "Testing", "High", "Apr 02",
            "Test core fastAPI endpoints",
            [
                "☐ Write unit tests for /predict/los and /predict/demand",
                "☐ Test 422 validations for bad inputs"
            ]
        ),
        (
            "Standardise API response casing",
            "Backend Dev", "Backend", "Medium", "Apr 01",
            "Convert to camelCase",
            [
                "☐ Apply generic camelCase alias_generator in pydantic",
                "☐ Update frontend fetch expectations"
            ]
        ),
        (
            "Update README for ML pipeline setup",
            "ML Engineer", "Documentation", "Medium", "Apr 04",
            "Document environment instructions",
            [
                "☐ Add steps for Python venv, requirements, node setup",
                "☐ Document env vars needed for supabase and models"
            ]
        )
    ],

    "In Progress": [
        (
            "ML Metrics dashboard polishing",
            "Frontend Dev", "Frontend", "High", "Mar 29",
            "Improve K-Fold comparison charts UI",
            [
                "☐ Add dropdown to select model (XGBoost/RF)",
                "☐ Add Recharts tooltip for exact metric values"
            ]
        ),
        (
            "Performance profiling training scripts",
            "ML Engineer", "ML/AI", "Medium", "Mar 30",
            "Speed up cross-validation",
            [
                "☐ Implement joblib Parallel(n_jobs=-1) in los_train.py",
                "☐ Document speedup results"
            ]
        ),
        (
            "PatientPrediction.jsx UX improvements",
            "Frontend Dev", "Frontend", "Medium", "Mar 30",
            "Smooth transition of results",
            [
                "☐ Implement skeleton loader during fetch",
                "☐ Add form validation feedback"
            ]
        )
    ],

    "Testing/Completed": [
        (
            "Supabase Auth Integration",
            "Full Stack", "Security", "High", "Mar 13",
            "Login persistence across system",
            [
                "✅ Session persisted via AuthContext & localStorage"
            ]
        ),
        (
            "Database Schema & RLS",
            "Backend Dev", "Database", "High", "Mar 13",
            "Setup patients, rooms, equipment tables",
            [
                "✅ Tables created and schema committed out to SQL",
                "✅ Basic RLS configurations validated"
            ]
        ),
        (
            "Core Frontend CRUD Modules",
            "Frontend Dev", "Frontend", "High", "Mar 18",
            "Pages for Inventory, Patients, Rooms",
            [
                "✅ Search & filter capabilities enabled",
                "✅ Sub-components implemented for modular UI"
            ]
        ),
        (
            "ML Predictor Models Trained",
            "ML Engineer", "ML/AI", "High", "Mar 22",
            "LOS, Equipment Failure, Medicine Demand",
            [
                "✅ XGBoost & Prophets models persisted to .pkl",
                "✅ Model metrics extracted and stored in JSON"
            ]
        ),
        (
            "Vision Anomaly Models",
            "ML Engineer", "ML/AI", "High", "Mar 24",
            "CNN Autoencoder and HF pipeline",
            [
                "✅ MRI Image processing pipeline integrated",
                "✅ Brain tumor pipeline dynamically served"
            ]
        ),
        (
            "Parallel API refactor",
            "Frontend Dev", "Performance", "High", "Mar 27",
            "Load time enhancement",
            [
                "✅ Converted sequential API grabs to Promise.all()"
            ]
        )
    ]
}

# ─────────────────────────────────────────────────────────────────────────────
# BUILD SHEET
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "MS Planner — HMS"
ws.sheet_view.showGridLines = False

# Column layout: bucket | # | task | assignee | label | priority | due | notes | checklist
COL_W = [24, 4, 42, 16, 16, 12, 10, 36, 60]
COLS  = ["Bucket", "#", "Task Name", "Assignee", "Label", "Priority", "Due", "Notes", "Checklist"]
for i, w in enumerate(COL_W, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:I1")
t = ws["A1"]
t.value = "🏥  MS Planner — AI-Powered Hospital Management System"
t.font = Font(bold=True, size=16, color="FFFFFF", name="Segoe UI")
t.alignment = Alignment(horizontal="center", vertical="center")
t.fill = fill(DARK_BG)
ws.row_dimensions[1].height = 36

# ── Column headers ─────────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 26
for col, header in enumerate(COLS, 1):
    c = ws.cell(row=2, column=col, value=header)
    c.font = Font(bold=True, size=10, color="FFFFFF", name="Segoe UI")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = fill("2E3250")
    c.border = border("1A1A2E")

# ── Write each bucket ─────────────────────────────────────────────────────────
current_row = 3
ALT = ["FFFFFF", "F8F9FF"]

for bucket_name, tasks in BUCKETS.items():
    accent, light_bg = BUCKET_CLR.get(bucket_name, ("999999", "EEEEEE"))

    # Bucket header row
    ws.merge_cells(f"A{current_row}:I{current_row}")
    bh = ws[f"A{current_row}"]
    bh.value = f"  {bucket_name.upper()}  ·  {len(tasks)} major tasks"
    bh.font = Font(bold=True, size=12, color="FFFFFF", name="Segoe UI")
    bh.alignment = Alignment(horizontal="left", vertical="center")
    bh.fill = fill(accent)
    bh.border = border(accent)
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Task rows
    for idx, (task, assignee, label, priority, due, notes, checklist) in enumerate(tasks):
        row = current_row
        alt_bg = ALT[idx % 2]

        ws.row_dimensions[row].height = max(45, len(checklist) * 16 + 8)

        # Col A
        bc = ws.cell(row=row, column=1, value=bucket_name)
        bc.font = Font(bold=True, size=8, color=accent, name="Segoe UI")
        bc.alignment = Alignment(horizontal="center", vertical="center")
        bc.fill = fill(light_bg)
        bc.border = border(accent)

        # Col B
        ic = ws.cell(row=row, column=2, value=idx + 1)
        ic.font = Font(size=9, color="888888", name="Segoe UI")
        ic.alignment = Alignment(horizontal="center", vertical="center")
        ic.fill = fill(alt_bg)
        ic.border = border()

        # Col C
        tc = ws.cell(row=row, column=3, value=task)
        tc.font = Font(bold=True, size=10, color="1A1A2E", name="Segoe UI")
        tc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        tc.fill = fill(alt_bg)
        tc.border = border()

        # Col D
        ac = ws.cell(row=row, column=4, value=assignee)
        ac.font = Font(size=9, italic=True, color="444466", name="Segoe UI")
        ac.alignment = Alignment(horizontal="center", vertical="center")
        ac.fill = fill(alt_bg)
        ac.border = border()

        # Col E
        lbg, lfc = LABEL_COLORS.get(label, ("EEEEEE", "333333"))
        lc = ws.cell(row=row, column=5, value=label)
        lc.font = Font(bold=True, size=9, color=lfc, name="Segoe UI")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = fill(lbg)
        lc.border = border()

        # Col F
        pbg, pfc = PRIORITY.get(priority, ("EEEEEE", "333333"))
        pc = ws.cell(row=row, column=6, value=priority)
        pc.font = Font(bold=True, size=9, color=pfc, name="Segoe UI")
        pc.alignment = Alignment(horizontal="center", vertical="center")
        pc.fill = fill(pbg)
        pc.border = border()

        # Col G
        dc = ws.cell(row=row, column=7, value=due)
        dc.font = Font(size=9, color="555555", name="Segoe UI")
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.fill = fill(alt_bg)
        dc.border = border()

        # Col H
        nc = ws.cell(row=row, column=8, value=notes)
        nc.font = Font(size=9, italic=True, color="555577", name="Segoe UI")
        nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        nc.fill = fill(alt_bg)
        nc.border = border()

        # Col I (checklist)
        checklist_text = "\n".join(checklist)
        cc = ws.cell(row=row, column=9, value=checklist_text)
        cc.font = Font(size=9, color="2C2C44", name="Segoe UI")
        cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cc.fill = fill(light_bg)
        cc.border = border(accent)

        current_row += 1

    # Spacer row
    ws.row_dimensions[current_row].height = 10
    for col in range(1, 10):
        ws.cell(current_row, col).fill = fill("F0F0F8")
        ws.cell(current_row, col).border = border("DADAEA")
    current_row += 1

# ── Summary row ────────────────────────────────────────────────────────────────
total = sum(len(t) for t in BUCKETS.values())
total_checklist = sum(len(t[6]) for tasks in BUCKETS.values() for t in tasks)
ws.merge_cells(f"A{current_row}:I{current_row}")
sr = ws[f"A{current_row}"]
sr.value = (
    f"  TOTAL MAJOR TASKS: {total}  ·  "
    f"TOTAL CHECKLIST ITEMS: {total_checklist}"
)
sr.font = Font(bold=True, size=10, color="FFFFFF", name="Segoe UI")
sr.alignment = Alignment(horizontal="center", vertical="center")
sr.fill = fill(DARK_BG)
sr.border = border(DARK_BG)
ws.row_dimensions[current_row].height = 24

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — Legend
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Legend")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:E1")
lt = ws2["A1"]
lt.value = "Legend"
lt.font = Font(bold=True, size=14, color="FFFFFF", name="Segoe UI")
lt.alignment = Alignment(horizontal="center", vertical="center")
lt.fill = fill(DARK_BG)

ws2.merge_cells("A3:B3")
sc = ws2["A3"]
sc.value = "☐ Open task"
ws2.merge_cells("A4:B4")
cc = ws2["A4"]
cc.value = "✅ Completed"

# ─────────────────────────────────────────────────────────────────────────────
output = r"d:\Minor Project goated\MS_Planner_Buckets.xlsx"
wb.save(output)

print(f"✅ Saved clean version: {output}")
for name, tasks in BUCKETS.items():
    print(f"   · {name:30s} {len(tasks)} major tasks")
